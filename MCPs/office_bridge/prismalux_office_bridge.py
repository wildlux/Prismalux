#!/usr/bin/env python3
# prismalux_office_bridge.py — Bridge HTTP per LibreOffice (UNO + file-based)
# ─────────────────────────────────────────────────────────────────────────────
# Modalità UNO (preferita): controllo diretto di LibreOffice tramite API UNO.
#   Requisiti: sudo apt install python3-uno  (Debian/Ubuntu/Mint)
#              LibreOffice installato (soffice nel PATH)
#
# Modalità file (fallback): crea file .docx/.xlsx/.pptx con python-docx/openpyxl/pptx.
#   Requisiti: pip install python-docx openpyxl python-pptx
#
# Avvio:
#   python3 prismalux_office_bridge.py
# ─────────────────────────────────────────────────────────────────────────────

import importlib
import io
import json
import os
import secrets
import subprocess
import sys
import threading
import time
import traceback
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path

PORT      = 6790
LO_PORT   = 2002
LO_ACCEPT = f"socket,host=localhost,port={LO_PORT};urp;StarOffice.ServiceManager"
LO_URL    = f"uno:{LO_ACCEPT}"

# ── Token di autenticazione ──────────────────────────────────────────
# Generato all'avvio, scritto in ~/.prismalux_office_token (chmod 600).
# Qt lo legge da disco e lo invia come "Authorization: Bearer <token>".
# Chiude il buco ACAO:* — senza il token nessuna pagina web può invocare
# il bridge anche se ascolta su localhost.
TOKEN_FILE = Path.home() / ".prismalux_office_token"
_token: str = ""

def _generate_token() -> str:
    global _token
    _token = secrets.token_hex(32)
    TOKEN_FILE.write_text(_token, encoding="utf-8")
    TOKEN_FILE.chmod(0o600)   # solo il proprietario può leggere
    return _token

def _verify_token(handler) -> bool:
    """Verifica l'header Authorization: Bearer <token>."""
    auth = handler.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return False
    return secrets.compare_digest(auth[7:], _token)

# ── Stato globale UNO ────────────────────────────────────────────────
_mode        = "file"   # "uno" | "file"
_desktop     = None     # com.sun.star.frame.Desktop
_smgr        = None     # com.sun.star.lang.XMultiServiceFactory
_lo_proc     = None     # subprocess.Popen di soffice
_uno_lock    = threading.Lock()   # serializza chiamate UNO


# ══════════════════════════════════════════════════════════════════════
# UNO — avvio e connessione
# ══════════════════════════════════════════════════════════════════════

def _uno_available() -> bool:
    try:
        import uno   # noqa: F401
        return True
    except ImportError:
        return False


def _find_soffice() -> str:
    """Cerca il binario soffice in PATH e nei percorsi comuni."""
    import shutil
    found = shutil.which("soffice")
    if found:
        return found
    candidates = [
        "/usr/bin/soffice",
        "/usr/lib/libreoffice/program/soffice",
        "/opt/libreoffice/program/soffice",
        "/snap/bin/soffice",                                  # Ubuntu snap
        "/var/lib/flatpak/exports/bin/org.libreoffice.LibreOffice",  # Flatpak
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",      # macOS
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for c in candidates:
        if Path(c).exists():
            return c
    return ""


def _start_lo_process(soffice: str) -> bool:
    """Avvia LibreOffice in modalità headless con socket UNO."""
    global _lo_proc
    # Non avviare se è già in esecuzione (connessione esterna)
    if _lo_proc and _lo_proc.poll() is None:
        return True
    try:
        _lo_proc = subprocess.Popen(
            [soffice, "--headless", "--norestore", "--nofirststartwizard",
             f"--accept={LO_ACCEPT}"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        return True
    except Exception as e:
        print(f"[Bridge] Errore avvio soffice: {e}")
        return False


def _connect_uno(max_tries: int = 15) -> bool:
    """Tenta la connessione UNO a LibreOffice (retry con backoff)."""
    global _desktop, _smgr
    import uno
    localContext = uno.getComponentContext()
    resolver = localContext.ServiceManager.createInstanceWithContext(
        "com.sun.star.bridge.UnoUrlResolver", localContext)
    for attempt in range(max_tries):
        try:
            ctx = resolver.resolve(LO_URL)
            _smgr = ctx.ServiceManager
            _desktop = _smgr.createInstanceWithContext(
                "com.sun.star.frame.Desktop", ctx)
            return True
        except Exception:
            time.sleep(1.0)
    return False


def _init_uno_mode() -> bool:
    """Tenta il setup completo UNO (avvio LO + connessione)."""
    global _mode
    if not _uno_available():
        print("[Bridge] python3-uno non trovato → modalità file")
        print("         Per UNO: sudo apt install python3-uno libreoffice")
        return False

    soffice = _find_soffice()
    if not soffice:
        print("[Bridge] soffice non trovato → modalità file")
        print("         Installa LibreOffice: sudo apt install libreoffice")
        return False

    print(f"[Bridge] Avvio LibreOffice UNO: {soffice}")
    if not _start_lo_process(soffice):
        return False

    print("[Bridge] Connessione UNO (attendo fino a 15s)...")
    if not _connect_uno(max_tries=15):
        print("[Bridge] Connessione UNO fallita → modalità file")
        return False

    _mode = "uno"
    print("[Bridge] Modalità UNO attiva — LibreOffice connesso")
    return True


# ══════════════════════════════════════════════════════════════════════
# Namespace per exec()
# ══════════════════════════════════════════════════════════════════════

def _make_namespace() -> dict:
    ns: dict = {"__builtins__": __builtins__, "Path": Path}

    # ── Moduli standard ──
    for name in ("os", "shutil", "datetime", "re", "math"):
        try:
            ns[name] = importlib.import_module(name)
        except ImportError:
            pass

    # ── UNO (se connesso) ──
    if _mode == "uno" and _desktop is not None:
        try:
            import uno
            from com.sun.star.beans import PropertyValue

            def createUnoService(name: str):
                return _smgr.createInstance(name)

            def systemPath(p: str) -> str:
                return uno.systemPathToFileUrl(str(p))

            def mkprops(**kwargs) -> tuple:
                result = []
                for k, v in kwargs.items():
                    pv = PropertyValue()
                    pv.Name = k
                    pv.Value = v
                    result.append(pv)
                return tuple(result)

            ns.update({
                "uno":              uno,
                "desktop":          _desktop,
                "PropertyValue":    PropertyValue,
                "createUnoService": createUnoService,
                "systemPath":       systemPath,   # systemPath(path) → "file:///..."
                "mkprops":          mkprops,      # mkprops(FilterName="writer8")
            })
        except Exception as e:
            print(f"[Bridge] Errore costruzione namespace UNO: {e}")

    # ── Librerie file-based (sempre disponibili come fallback) ──
    try:
        from docx import Document
        from docx.shared import Pt, Cm, RGBColor
        from docx.enum.text import WD_ALIGN_PARAGRAPH
        ns.update({"Document": Document, "Pt": Pt, "Cm": Cm,
                   "RGBColor": RGBColor, "WD_ALIGN_PARAGRAPH": WD_ALIGN_PARAGRAPH})
    except ImportError:
        pass
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference
        ns.update({"Workbook": Workbook, "load_workbook": load_workbook,
                   "Font": Font, "PatternFill": PatternFill,
                   "Alignment": Alignment, "Border": Border, "Side": Side,
                   "BarChart": BarChart, "Reference": Reference})
    except ImportError:
        pass
    try:
        from pptx import Presentation
        from pptx.util import Inches, Pt as PptPt
        from pptx.dml.color import RGBColor as PptRGB
        ns.update({"Presentation": Presentation, "Inches": Inches,
                   "PptPt": PptPt, "PptRGB": PptRGB})
    except ImportError:
        pass

    return ns


def _check_libraries() -> dict:
    libs = {}
    for lib, mod in [("python-docx", "docx"),
                     ("openpyxl",    "openpyxl"),
                     ("python-pptx", "pptx"),
                     ("python3-uno", "uno")]:
        try:
            importlib.import_module(mod)
            libs[lib] = True
        except ImportError:
            libs[lib] = False
    return libs


# ══════════════════════════════════════════════════════════════════════
# HTTP Handler
# ══════════════════════════════════════════════════════════════════════

class _Handler(BaseHTTPRequestHandler):

    def log_message(self, fmt, *args):
        pass

    def _send_json(self, code: int, obj: dict):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type",   "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        # ACAO limitato a localhost — nessuna pagina web esterna può invocare il bridge
        self.send_header("Access-Control-Allow-Origin", "http://localhost")
        self.end_headers()
        self.wfile.write(body)

    def _unauthorized(self):
        self._send_json(401, {"ok": False, "error": "Non autorizzato — token mancante o errato"})

    def do_GET(self):
        # /status richiede token: anche il polling di stato non deve essere pubblico
        if not _verify_token(self):
            self._unauthorized()
            return
        if self.path == "/status":
            lo_running = _lo_proc is not None and _lo_proc.poll() is None
            self._send_json(200, {
                "ok":         True,
                "mode":       _mode,
                "uno":        _mode == "uno",
                "lo_running": lo_running,
                "libraries":  _check_libraries(),
                "port":       PORT,
            })
        else:
            self._send_json(404, {"ok": False, "error": "endpoint non trovato"})

    def do_POST(self):
        if not _verify_token(self):
            self._unauthorized()
            return

        if self.path != "/execute":
            self._send_json(404, {"ok": False, "error": "endpoint non trovato"})
            return

        length = int(self.headers.get("Content-Length", 0))
        raw    = self.rfile.read(length)
        try:
            data = json.loads(raw)
            code = data.get("code", "").strip()
        except Exception as e:
            self._send_json(400, {"ok": False, "error": f"JSON non valido: {e}"})
            return

        if not code:
            self._send_json(400, {"ok": False, "error": "campo 'code' vuoto"})
            return

        buf     = io.StringIO()
        old_out = sys.stdout
        try:
            sys.stdout = buf
            with _uno_lock:
                ns = _make_namespace()
                exec(compile(code, "<prismalux_office>", "exec"), ns)
            sys.stdout = old_out
            output = buf.getvalue().strip()
            self._send_json(200, {
                "ok":     True,
                "mode":   _mode,
                "output": output or "Eseguito con successo.",
            })
        except Exception:
            sys.stdout = old_out
            self._send_json(200, {
                "ok":    False,
                "mode":  _mode,
                "error": traceback.format_exc().strip(),
            })


# ══════════════════════════════════════════════════════════════════════
# Avvio
# ══════════════════════════════════════════════════════════════════════

def main():
    # Token generato all'avvio — scritto in ~/.prismalux_office_token
    tok = _generate_token()
    print(f"\n{'='*60}")
    print( "  Prismalux Office Bridge")
    print(f"  Porta : {PORT}")
    print(f"  Token : {TOKEN_FILE}")
    print(f"{'='*60}")

    # Tenta modalità UNO; in caso di fallimento resta in file mode
    _init_uno_mode()

    if _mode == "file":
        libs = _check_libraries()
        ok   = [k for k, v in libs.items() if v and k != "python3-uno"]
        miss = [k for k, v in libs.items() if not v and k != "python3-uno"]
        print(f"\n  Modalità FILE")
        if ok:   print(f"  Disponibili: {', '.join(ok)}")
        if miss: print(f"  Mancanti (pip install): {', '.join(miss)}")

    print(f"\n  Endpoint: http://localhost:{PORT}/execute")
    print("  Premi Ctrl+C per fermare.\n")

    server = HTTPServer(("localhost", PORT), _Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[Bridge] Arresto...")
    finally:
        server.server_close()
        # Rimuove il token file all'uscita
        try:
            TOKEN_FILE.unlink(missing_ok=True)
        except Exception:
            pass
        if _lo_proc and _lo_proc.poll() is None:
            print("[Bridge] Chiusura LibreOffice...")
            _lo_proc.terminate()
            _lo_proc.wait(timeout=5)
        print("[Bridge] Fermato.")


if __name__ == "__main__":
    main()
